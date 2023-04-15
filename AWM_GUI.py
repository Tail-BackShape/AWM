import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.messagebox as mb
import openpyxl as pyxl
from openpyxl.styles import Font

font = Font(name="游ゴシック")
max_r, max_c = 0, 0
sh, newsh = None, None

def open_file():
    filepath = filedialog.askopenfilename(
        initialdir="./",
        title="Select File",
        filetypes=[("Excel Files", "*.xl*")]
    )
    filepath_entry.delete(0, tk.END)
    filepath_entry.insert(tk.END, filepath)

    global sh, max_r, max_c
    wb = pyxl.load_workbook(filepath)
    sh = wb["Sheet1"]
    max_r = sh.max_row
    max_c = sh.max_column

def make_wing():
    wing_span = int(wing_span_entry.get())

    global newsh
    newwb = pyxl.Workbook() #新規ファイル作成
    newsh = newwb.active #アクティブシート読み込み

    for i in range(max_r):
        i += 1
        for j in range(max_c):
            j += 1
            data = sh.cell(i, j).value

            if not(data is None):
                c_data = data * wing_span
                cell = newsh.cell(i, j)
                cell.value = (c_data)
                cell.font = font

    filepath = filedialog.asksaveasfilename(
        initialdir="./",
        title="Save File",
        filetypes=[("Excel Files", "*.xlsx*")],
        defaultextension=".xlsx"
    )

    if filepath:
        newwb.save(filepath)
        mb.showinfo("Result","Process completed!")
        wing_span_entry.delete(0, tk.END)
    else:
        mb.showinfo("Result","Process aborted.")

root = tk.Tk()
root.title("Auto Wing Maker")

frame = ttk.Frame(root, padding=10)
frame.grid()

filepath_label = ttk.Label(frame, text="Reference data file:")
filepath_label.grid(column=0, row=0, sticky=tk.W)

filepath_entry = ttk.Entry(frame, width=50)
filepath_entry.grid(column=0, row=1, sticky=tk.W)

filepath_button = ttk.Button(frame, text="Open", command=open_file)
filepath_button.grid(column=1, row=1, sticky=tk.W)

wing_span_label = ttk.Label(frame, text="Wing span (mm):")
wing_span_label.grid(column=0, row=4, sticky=tk.W)

wing_span_entry = ttk.Entry(frame, width=10)
wing_span_entry.grid(column=0, row=5, sticky=tk.W)

calculate_button = ttk.Button(frame, text="Save", command=make_wing)
calculate_button.grid(column=1, row=5, sticky=tk.W)

root.mainloop()
